"""
Generate an Excel workbook for manual furnishing review and override.

POLYGON-FIRST: Only polygons from polygons.json are listed.
Polygon primary_function and area_m2 are the authoritative values.
DB Space data is shown for reference only — it never overrides polygon data.

Sheets:
  1. All Spaces        – every polygon with current furnishings, rule match, and override columns
  2. Furnishing Rules  – reference: all current rules with keywords and default furnishings
  3. Furnishing Catalog – reference: all item types with footprint and occupancy values
  4. Excluded Patterns  – reference: patterns that block furnishing assignment

Run from backend/:
    python generate_furnishing_excel.py
"""

import io
import json
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

# Fix Windows encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# Setup path so app imports work
sys.path.insert(0, str(Path(__file__).resolve().parent))
os.chdir(Path(__file__).resolve().parent)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import DATA_DIR
from app.database import SessionLocal
from app.models import Space, SpaceMetrics, SpaceFurnishing, FurnishingType
from app.services.furnishings import (
    FURNISHING_CATALOG,
    FUNCTION_FURNISHING_RULES,
    EXCLUDED_PATTERNS,
    MAX_FURNISHING_PCT,
    _is_excluded,
    _matches_any,
    _apply_scaling,
    _compute_total_footprint,
)

POLYGONS_FILE = DATA_DIR / "polygons.json"


def read_polygons():
    if not POLYGONS_FILE.exists():
        return []
    with open(POLYGONS_FILE, "r", encoding="utf-8") as f:
        content = f.read()
    if not content.strip():
        return []
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        cleaned = re.sub(r',\s*([}\]])', r'\1', content)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            decoder = json.JSONDecoder()
            result, _ = decoder.raw_decode(cleaned.lstrip())
            return result if isinstance(result, list) else []


def determine_rule_match(fn: str, area_m2: float, ft_footprints: dict):
    """Simulate the seeder's rule matching. Returns (status, rule_index, rule_keywords, furnishing_list)."""
    if not fn or fn == "?" or fn == "Unassigned":
        return ("no_function", None, None, None)

    if _is_excluded(fn):
        return ("excluded", None, None, None)

    for idx, (keywords, base_furnishings, options) in enumerate(FUNCTION_FURNISHING_RULES):
        if _matches_any(fn, keywords):
            min_area = options.get("min_area", 0)
            if area_m2 > 0 and area_m2 < min_area:
                return ("too_small", idx, keywords[0], None)

            scale_rules = options.get("scale", {})
            if scale_rules and area_m2 > 0:
                furnishing_list = _apply_scaling(base_furnishings, area_m2, scale_rules, ft_footprints)
            else:
                furnishing_list = list(base_furnishings)

            if area_m2 > 0:
                total_fp = _compute_total_footprint(furnishing_list, ft_footprints)
                if total_fp > area_m2 * MAX_FURNISHING_PCT:
                    return ("overflow", idx, keywords[0], furnishing_list)

            return ("matched", idx, keywords[0], furnishing_list)

    return ("no_rule_match", None, None, None)


def main():
    db = SessionLocal()

    # --- Load all data ---
    print("Loading data...")
    polygons = read_polygons()
    polygons = [p for p in polygons if p.get("ifc_guid")]
    print(f"Polygons loaded: {len(polygons)}")

    # DB spaces — reference only, never overrides polygon data
    db_spaces = db.query(Space).filter(Space.ifc_guid.isnot(None)).all()
    db_space_map = {sp.ifc_guid: sp for sp in db_spaces}

    all_metrics = {m.ifc_guid: m for m in db.query(SpaceMetrics).all()}
    all_furnishings = defaultdict(list)
    for f in db.query(SpaceFurnishing).all():
        all_furnishings[f.ifc_guid].append(f)

    ft_map = {ft.item_type: ft for ft in db.query(FurnishingType).all()}
    ft_footprints = {ft.item_type: ft.footprint_m2 for ft in ft_map.values()}

    # --- Create workbook ---
    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # SHEET 1: All Spaces (polygon-driven)
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "All Spaces"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    override_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    override_header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    override_header_font = Font(bold=True, color="000000", size=11)
    ref_header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ref_header_font = Font(bold=True, color="1F3864", size=11)
    ref_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
    status_colors = {
        "matched":       PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "excluded":      PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "no_function":   PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "too_small":     PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "overflow":      PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "no_rule_match": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        # --- Polygon data (authoritative) ---
        "Floor ID",                     # A
        "IFC GUID",                     # B
        "Polygon Space Name",           # C
        "Polygon Primary Function",     # D  — THIS is the effective function
        "Polygon Area (m\u00b2)",       # E  — THIS is the effective area
        # --- Rule matching ---
        "Rule Status",                  # F
        "Matched Rule #",              # G
        "Matched Keyword",             # H
        # --- Current furnishings ---
        "Current Furnishings",          # I
        "Total Footprint (m\u00b2)",    # J
        "Footprint %",                  # K
        "Normal Occ",                   # L
        "Max Occ",                      # M
        "Absolute Occ",                # N
        "Used Area (m\u00b2)",          # O
        "Free Area (m\u00b2)",          # P
        "Furnishing Source",            # Q
        # --- DB reference (light blue, informational only) ---
        "DB Space Name (ref)",          # R
        "DB Primary Function (ref)",    # S
        "DB Area m\u00b2 (ref)",        # T
        "Has DB Record",                # U
        # --- Override columns (yellow) ---
        "OVERRIDE: Rule #",             # V
        "OVERRIDE: Furnishings",        # W  — e.g. "patient_bed:2, desk:1"
        "OVERRIDE: Notes",              # X
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_idx <= 17:  # Polygon + rule + furnishing columns
            cell.font = header_font
            cell.fill = header_fill
        elif col_idx <= 21:  # DB reference columns
            cell.font = ref_header_font
            cell.fill = ref_header_fill
        else:  # Override columns
            cell.font = override_header_font
            cell.fill = override_header_fill

    # Freeze header + first 2 ID columns
    ws.freeze_panes = "C2"

    # --- Populate rows (polygon-driven only) ---
    rows_data = []
    for poly in polygons:
        guid = poly["ifc_guid"]
        floor_id = poly.get("floor_id", "")
        space_name = poly.get("space_name", "")
        poly_function = poly.get("primary_function", "") or ""
        area_m2 = poly.get("area_m2") or 0

        # DB space — reference only
        sp = db_space_map.get(guid)
        db_name = sp.space_name if sp else ""
        db_function = sp.primary_function if sp else ""
        db_area = sp.area_m2 if sp else ""
        has_db = "Yes" if sp else "No"

        # Rule matching against POLYGON function (not DB)
        status, rule_idx, matched_kw, furn_list = determine_rule_match(
            poly_function, area_m2 or 0, ft_footprints)

        # Current furnishings from DB
        curr_furnishings = all_furnishings.get(guid, [])
        curr_furn_str = ", ".join(
            f"{f.item_type} x{f.quantity}" for f in sorted(curr_furnishings, key=lambda x: x.item_type)
        ) if curr_furnishings else ""

        # Total footprint of current furnishings
        total_fp = sum(ft_footprints.get(f.item_type, 0) * f.quantity for f in curr_furnishings)
        fp_pct = round(total_fp / area_m2 * 100, 1) if area_m2 > 0 else 0

        # Metrics
        m = all_metrics.get(guid)

        rows_data.append({
            "floor_id": floor_id,
            "guid": guid,
            "space_name": space_name,
            "poly_function": poly_function,
            "area_m2": round(area_m2, 2) if area_m2 else "",
            "status": status,
            "rule_idx": rule_idx if rule_idx is not None else "",
            "matched_kw": matched_kw or "",
            "curr_furn_str": curr_furn_str,
            "total_fp": round(total_fp, 2) if total_fp > 0 else "",
            "fp_pct": fp_pct if fp_pct > 0 else "",
            "normal_occ": m.normal_occupancy if m else "",
            "max_occ": m.max_occupancy if m else "",
            "abs_occ": m.absolute_occupancy if m else "",
            "used_area": m.used_area_m2 if m and m.used_area_m2 is not None else "",
            "free_area": m.free_area_m2 if m and m.free_area_m2 is not None else "",
            "furn_source": m.furnishing_source if m else "",
            "db_name": db_name or "",
            "db_function": db_function or "",
            "db_area": round(db_area, 2) if db_area else "",
            "has_db": has_db,
        })

    # Sort by floor, then status priority, then function
    status_order = {"no_rule_match": 0, "overflow": 1, "too_small": 2, "no_function": 3, "matched": 4, "excluded": 5}
    rows_data.sort(key=lambda r: (r["floor_id"], status_order.get(r["status"], 9), r["poly_function"]))

    for row_idx, r in enumerate(rows_data, 2):
        values = [
            r["floor_id"], r["guid"], r["space_name"],
            r["poly_function"], r["area_m2"],
            r["status"], r["rule_idx"], r["matched_kw"],
            r["curr_furn_str"], r["total_fp"], r["fp_pct"],
            r["normal_occ"], r["max_occ"], r["abs_occ"],
            r["used_area"], r["free_area"], r["furn_source"],
            r["db_name"], r["db_function"], r["db_area"], r["has_db"],
            "", "", "",  # Override columns
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # Color status column (F = col 6)
            if col_idx == 6:
                fill = status_colors.get(val)
                if fill:
                    cell.fill = fill
            # Light blue for DB reference columns (R-U = 18-21)
            if 18 <= col_idx <= 21:
                cell.fill = ref_fill
            # Yellow for override columns (V-X = 22-24)
            if col_idx >= 22:
                cell.fill = override_fill

    # Column widths
    col_widths = {
        1: 8, 2: 14, 3: 25, 4: 45, 5: 12,
        6: 14, 7: 10, 8: 22,
        9: 55, 10: 12, 11: 10, 12: 10, 13: 10, 14: 10, 15: 12, 16: 12, 17: 14,
        18: 25, 19: 40, 20: 12, 21: 10,
        22: 14, 23: 55, 24: 30,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows_data)+1}"

    # ═══════════════════════════════════════════════════════════════
    # SHEET 2: Furnishing Rules (reference)
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Furnishing Rules")
    rule_headers = ["Rule #", "Keywords", "Default Furnishings", "Min Area (m\u00b2)",
                    "Scaling Rules", "Total Base Footprint (m\u00b2)"]

    for col_idx, h in enumerate(rule_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for idx, (keywords, base_furnishings, options) in enumerate(FUNCTION_FURNISHING_RULES):
        row = idx + 2
        kw_str = ", ".join(keywords)
        furn_str = ", ".join(f"{it} x{q}" for it, q in base_furnishings)
        min_area = options.get("min_area", "")
        scale = options.get("scale", {})
        scale_str = ", ".join(
            f"{k}: 1 per {v['per_m2']}m\u00b2 (min {v['min']}, max {v['max']})"
            for k, v in scale.items()
        ) if scale else ""
        base_fp = sum(ft_footprints.get(it, 0) * q for it, q in base_furnishings)

        vals = [idx, kw_str, furn_str, min_area, scale_str, round(base_fp, 2)]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 45
    ws2.column_dimensions["F"].width = 18
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    # SHEET 3: Furnishing Catalog (reference)
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Furnishing Catalog")
    cat_headers = ["Item Type", "Category", "Label", "Footprint (m\u00b2)", "Normal Occ", "Max Occ"]

    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for idx, (item_type, category, label, footprint, normal_occ, max_occ) in enumerate(FURNISHING_CATALOG):
        row = idx + 2
        vals = [item_type, category, label, footprint, normal_occ, max_occ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 10
    ws3.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    # SHEET 4: Excluded Patterns (reference)
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Excluded Patterns")
    cell = ws4.cell(row=1, column=1, value="Pattern")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell = ws4.cell(row=1, column=2, value="Polygons Matching")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

    for idx, pattern in enumerate(EXCLUDED_PATTERNS):
        row = idx + 2
        count = sum(1 for r in rows_data if r["poly_function"] and pattern in r["poly_function"].lower())
        ws4.cell(row=row, column=1, value=pattern).border = thin_border
        ws4.cell(row=row, column=2, value=count).border = thin_border

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 18
    ws4.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════════
    # SHEET 5: Summary Stats
    # ═══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Summary")
    summary_data = defaultdict(int)
    floor_stats = defaultdict(lambda: defaultdict(int))
    for r in rows_data:
        summary_data[r["status"]] += 1
        floor_stats[r["floor_id"]][r["status"]] += 1

    cell = ws5.cell(row=1, column=1, value="Status")
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
    cell = ws5.cell(row=1, column=2, value="Count")
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border

    for idx, (status, count) in enumerate(sorted(summary_data.items(), key=lambda x: -x[1])):
        ws5.cell(row=idx+2, column=1, value=status).border = thin_border
        ws5.cell(row=idx+2, column=2, value=count).border = thin_border

    # Floor breakdown
    start_row = len(summary_data) + 4
    ws5.cell(row=start_row, column=1, value="Floor Breakdown").font = Font(bold=True, size=12)

    statuses = sorted(summary_data.keys())
    floor_headers = ["Floor"] + statuses + ["Total"]
    for col_idx, h in enumerate(floor_headers, 1):
        cell = ws5.cell(row=start_row+1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border

    for f_idx, floor_id in enumerate(sorted(floor_stats.keys())):
        row = start_row + 2 + f_idx
        ws5.cell(row=row, column=1, value=floor_id).border = thin_border
        total = 0
        for s_idx, status in enumerate(statuses):
            count = floor_stats[floor_id][status]
            total += count
            ws5.cell(row=row, column=s_idx+2, value=count).border = thin_border
        ws5.cell(row=row, column=len(statuses)+2, value=total).border = thin_border

    ws5.column_dimensions["A"].width = 18
    ws5.column_dimensions["B"].width = 12
    for i in range(len(statuses)):
        ws5.column_dimensions[get_column_letter(i+2)].width = 14

    # ═══════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════
    output_path = DATA_DIR / "furnishing_review.xlsx"
    wb.save(str(output_path))
    print(f"\nSaved to: {output_path}")

    # Print summary
    print(f"\nTotal polygons: {len(rows_data)}")
    for status, count in sorted(summary_data.items(), key=lambda x: -x[1]):
        print(f"  {status}: {count}")

    db.close()


if __name__ == "__main__":
    main()
