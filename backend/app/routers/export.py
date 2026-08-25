"""
Export endpoint: receives IFC metaScene dump from frontend,
combines with database spaces, and generates the rename registry Excel.
"""

from fastapi import APIRouter
from pydantic import BaseModel
from typing import Optional
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os

router = APIRouter()

DATA_DIR = Path(__file__).resolve().parent.parent.parent.parent / "data"
DB_PATH = DATA_DIR / "delta.db"
OUTPUT_PATH = DATA_DIR / "delta_rename_registry.xlsx"

# Debug: print resolved paths on import
print(f"[Export] DATA_DIR: {DATA_DIR}")
print(f"[Export] DB exists: {DB_PATH.exists()}")
print(f"[Export] Output will go to: {OUTPUT_PATH}")


class IfcElement(BaseModel):
    id: str
    type: str
    name: Optional[str] = None
    parent_name: Optional[str] = None
    parent_type: Optional[str] = None


class ExportRequest(BaseModel):
    elements: list[IfcElement]


# Shared styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2D3548", end_color="2D3548", fill_type="solid")
NEW_COL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
NEW_COL_FONT = Font(bold=True, color="E77133", size=11)
CELL_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(cell, is_new=False):
    cell.font = NEW_COL_FONT if is_new else HEADER_FONT
    cell.fill = NEW_COL_FILL if is_new else HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER


def auto_width(ws, columns, max_rows=100):
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(max_rows + 2, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 45))
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max_len + 3


@router.post("/api/export/rename-registry")
def generate_rename_registry(req: ExportRequest):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Spaces ──
    ws1 = wb.active
    ws1.title = "Spaces"

    space_cols = [
        "id", "ifc_guid", "space_name", "room_number",
        "floor_id", "section", "service_code",
        "primary_function", "secondary_functions", "space_class",
        "functional_zone", "area_m2", "subproject", "construction_phase",
        "New_Name", "New_Function",
    ]

    db_cols = [c for c in space_cols if c not in ("New_Name", "New_Function")]

    for col_idx, col_name in enumerate(space_cols, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, is_new=col_name.startswith("New_"))

    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute(f"SELECT {', '.join(db_cols)} FROM spaces ORDER BY floor_id, id")
    rows = cursor.fetchall()
    conn.close()

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
        # New_Name + New_Function columns (blank, highlighted)
        for offset in range(2):
            cell = ws1.cell(row=row_idx, column=len(db_cols) + 1 + offset, value="")
            cell.fill = CELL_FILL
            cell.border = THIN_BORDER

    auto_width(ws1, space_cols)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: IFC Elements (grouped by type + name) ──
    ws2 = wb.create_sheet("IFC_Elements")

    elem_cols = [
        "ifc_type", "ifc_name", "instance_count",
        "sample_guids", "parent_type", "parent_name",
        "New_Name", "New_Function",
    ]

    for col_idx, col_name in enumerate(elem_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, is_new=col_name.startswith("New_"))

    # Group elements by (type, name)
    groups = {}
    for el in req.elements:
        # Skip IfcSpace — those are in Sheet 1
        if el.type == "IfcSpace":
            continue
        key = (el.type, el.name or "")
        if key not in groups:
            groups[key] = {
                "guids": [],
                "parent_type": el.parent_type or "",
                "parent_name": el.parent_name or "",
            }
        groups[key]["guids"].append(el.id)

    # Sort by type then name
    sorted_groups = sorted(groups.items(), key=lambda x: (x[0][0], x[0][1]))

    for row_idx, ((ifc_type, ifc_name), info) in enumerate(sorted_groups, 2):
        # Show up to 3 sample GUIDs
        sample = ", ".join(info["guids"][:3])
        if len(info["guids"]) > 3:
            sample += " ..."

        values = [
            ifc_type,
            ifc_name,
            len(info["guids"]),
            sample,
            info["parent_type"],
            info["parent_name"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

        # New_Name + New_Function columns
        for offset in range(2):
            cell = ws2.cell(row=row_idx, column=len(values) + 1 + offset, value="")
            cell.fill = CELL_FILL
            cell.border = THIN_BORDER

    auto_width(ws2, elem_cols)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # Save
    save_path = str(OUTPUT_PATH)
    print(f"[Export] Saving to: {save_path}")
    try:
        wb.save(save_path)
        file_exists = os.path.exists(save_path)
        file_size = os.path.getsize(save_path) if file_exists else 0
        print(f"[Export] Saved OK. Exists: {file_exists}, Size: {file_size} bytes")
    except Exception as e:
        print(f"[Export] Save FAILED: {e}")
        return {"status": "error", "detail": str(e)}

    return {
        "status": "ok",
        "path": save_path,
        "file_size_kb": round(file_size / 1024, 1),
        "spaces_count": len(rows),
        "element_groups": len(sorted_groups),
        "total_elements": sum(len(g["guids"]) for g in groups.values()),
    }
